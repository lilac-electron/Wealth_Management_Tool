from flask import (
    Flask, render_template, jsonify, request, redirect, url_for, flash, session
)
import itsdangerous
import wtforms
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from wtforms import (
    StringField, SubmitField, IntegerField, PasswordField, BooleanField, SelectMultipleField,FileField
)
from wtforms.validators import (
    InputRequired, Email, Length, NumberRange, DataRequired, Optional#, validators
)
import email_validator

from wtforms import MultipleFileField
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, UserMixin, login_user, login_required, logout_user, current_user
)
from openpyxl import Workbook, load_workbook
import datetime
#from datetime import datetime
import pandas_datareader.data as pdr
import yfinance as yfin
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
import matplotlib.pyplot as plt
import os
import io
import base64
import csv 
import numpy as np
import requests
import json
from PythonGenerators.generateAccounts import FinanceDataGenerator
from PythonGenerators.generateCreditCardAccounts import CreditCardDataGenerator

#FinsFintechFYP

np.random.seed(seed=8)
app = Flask(__name__)

bootstrap = Bootstrap(app)

app.config['SECRET_KEY'] = 'your_secret_key'
app.config['WTF_CSRF_SECRET_KEY'] = 'a_random_key' 
#app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///PythonDB.db'
app.config['UPLOAD_FOLDER'] = 'upload_folder'

SQLALCHEMY_DATABASE_URI = "mysql+mysqlconnector://{username}:{password}@{hostname}/{databasename}".format(
    username="Finnnnnn1",
    password="FinsFintechFYP",
    hostname="Finnnnnn1.mysql.pythonanywhere-services.com",
    databasename="Finnnnnn1$users",
)
app.config["SQLALCHEMY_DATABASE_URI"] = SQLALCHEMY_DATABASE_URI
app.config["SQLALCHEMY_POOL_RECYCLE"] = 299
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False


db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(15), unique=True)
    email = db.Column(db.String(50), unique=True)
    password = db.Column(db.String(80))

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

class LoginForm(FlaskForm):
    username = StringField('Enter Username', validators=[InputRequired(), Length(min=3, max = 15)])
    password = PasswordField('Enter Password', validators=[InputRequired(), Length(min=8, max=80)])
    remember = BooleanField('Remember me')

class RegistrationForm(FlaskForm):
    email = StringField('Enter Email', validators=[InputRequired(), Email(message='Invalid Email'), Length(max=50)])
    username = StringField('Enter Username', validators=[InputRequired(), Length(min=3, max = 15)])
    password = PasswordField('Enter Password', validators=[InputRequired(), Length(min=8, max=80)])

#class MultipleFileUploadForm(FlaskForm):
#    files = MultipleFileField('Upload Files', validators=[FileAllowed(['csv', 'xls', 'xlsx'], 'Only CSV and Excel files are allowed.')])

class DeleteFileForm(FlaskForm):
    files_to_delete = SelectMultipleField('Files to Delete', coerce=str)
    submit = SubmitField('Delete Selected Files')

class DynamicForm(FlaskForm):
    pass
class DynamicForm2(FlaskForm):
    pass

class UploadForm(FlaskForm):
    file = FileField('Upload File', validators=[DataRequired()])
    submit = SubmitField('Submit')

class Transaction:
    def __init__(self, transaction_id, date, amount, description, category, balance, merchant=None):
        self.transaction_id = transaction_id
        self.date = date
        self.amount = amount
        self.description = description
        self.category = category
        self.merchant = merchant
        self.balance = balance

    def to_dict(self):
        transaction_dict = {
            "transaction_id": self.transaction_id,
            "date": self.date,
            "amount": self.amount,
            "description": self.description,
            "category": self.category,
            "balance": self.balance
        }
        if self.merchant:
            transaction_dict["merchant"] = self.merchant
        return transaction_dict

class Account:
    def __init__(self, account_number, sort_code, balance, currency, account_holder, transactions):
        self.account_number = account_number
        self.sort_code = sort_code
        self.balance = balance
        self.currency = currency
        self.account_holder = account_holder
        self.transactions = transactions
        self.account_total_in = 0
        self.account_total_out = 0 
        self.merchant_count = {}
        self.salary = 0
        for transaction in self.transactions:
            if transaction.amount > 0 :
                self.account_total_in += transaction.amount
            else:
                self.account_total_out += transaction.amount
                self.merchant_count[transaction.merchant['category']] = self.merchant_count.get(transaction.merchant['category'], 0)+1
            if transaction.description == "Salary Deposit":
                self.salary = float(transaction.amount)
        self.top_merchant = max(self.merchant_count, key=self.merchant_count.get)
        self.account_total_in = round(float(self.account_total_in), 2)
        self.account_total_out = round(float(self.account_total_out), 2)

    def to_dict(self):
        account_dict = {
            "account_number": self.account_number,
            "sort_code": self.sort_code,
            "balance": self.balance,
            "currency": self.currency,
            "account_holder": self.account_holder,
            "transactions": [txn.to_dict() for txn in self.transactions],
            "total_in": self.account_total_in,
            "total_out": self.account_total_out,
            "top_merchant": self.top_merchant,
            "salary": self.salary
        }
        return account_dict

class CreditCard:
    def __init__(self, card_number, expiry_date, card_holder, credit_limit, currency, transactions):
        self.card_number = card_number
        self.expiry_date = expiry_date
        self.card_holder = card_holder
        self.credit_limit = credit_limit
        self.currency = currency
        self.transactions = transactions
        self.credit_total_in = 0
        self.credit_total_out = 0 
        self.merchant_count = {}
        for transaction in self.transactions:
            if transaction.amount > 0 :
                self.credit_total_in += transaction.amount
            else:
                self.credit_total_out += transaction.amount
                self.merchant_count[transaction.merchant['category']] = self.merchant_count.get(transaction.merchant['category'], 0)+1
        self.top_merchant = max(self.merchant_count, key=self.merchant_count.get)
        self.credit_total_in = round(float(self.credit_total_in), 2)
        self.credit_total_out = round(float(self.credit_total_out), 2)
    

    def to_dict(self):
        credit_card_dict = {
            "card_number": self.card_number,
            "expiry_date": self.expiry_date,
            "card_holder": self.card_holder,
            "credit_limit": self.credit_limit,
            "currency": self.currency,
            "transactions": [txn.to_dict() for txn in self.transactions],
            "total_in": self.credit_total_in,
            "total_out": self.credit_total_out,
            "top_merchant": self.top_merchant
        }
        return credit_card_dict

class Form1(FlaskForm):
    field1 = StringField('Field 1', validators=[InputRequired()])

class Form2(FlaskForm):
    field2 = StringField('Field 2', validators=[InputRequired()])

def clearAttribute():
    DynamicForm = [attr for attr in DynamicForm if  not(attr.startswith('field_'))]

#def CreditsForm(inputs_list):
#    for item in inputs_list:
 #       setattr(DynamicForm, f'field_{item}', IntegerField(f'Enter monthly payment for {item}'))
        #print(item)
    #DynamicForm.submit = SubmitField('Submit')   

#def AssetForm(inputs_list):
 #   for item in inputs_list:
  #      setattr(DynamicForm2, f'field_{item}', IntegerField(f'Enter asset value for {item}'))
   #     #print(item)   

def list_files(directory):
    file_names = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]
    return file_names

def create_excel_file(file_path, sheet_data):
    # Create a Pandas Excel writer using xlsxwriter as the engine
    writer = pd.ExcelWriter(file_path)
    
    # Iterate through the sheet data
    for sheet_name, column_names in sheet_data.items():
        # Create a DataFrame with column names
        df = pd.DataFrame(columns=column_names)
        
        # Fill the DataFrame with zeros below the column names
        for col in column_names:
            df[col] = [0]
        
        # Write the DataFrame to the Excel file
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Save the Excel file
    writer.save()

def excel_to_dict(file_path):
    # Read the Excel file
    xls = pd.ExcelFile(file_path)
    
    # Initialize an empty dictionary to store sheet data
    excel_dict = {}
    
    # Iterate through each sheet in the Excel file
    for sheet_name in xls.sheet_names:
        # Read the sheet into a DataFrame
        df = pd.read_excel(xls, sheet_name)
        
        # Convert the DataFrame to a dictionary and add it to the excel_dict
        excel_dict[sheet_name] = df.to_dict(orient='records')
    
    return excel_dict

def write_dict_to_excel(file_path, sheet_name, data_dict):
    # Load the workbook or create a new one if it doesn't exist
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()
    
    # Select the worksheet or create a new one if it doesn't exist
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Write the keys in the first row
    keys = list(data_dict.keys())
    for col, key in enumerate(keys, start=1):
        ws.cell(row=1, column=col, value=key)

    # Write the values in the second row
    values = list(data_dict.values())
    for col, value in enumerate(values, start=1):
        ws.cell(row=2, column=col, value=value)

    # Save the workbook
    wb.save(file_path)

def create_csv_file(file_path, column_names, data_dict):
    with open(file_path, 'w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=column_names)
        writer.writeheader()
        writer.writerow(data_dict)

def read_csv_file(file_path):
    data_dict = {}
    with open(file_path, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            data_dict.update(row)
    return data_dict

def write_csv_file(upload_folder_path, data):
    # Ensure the folder exists (create it if it doesn't)
    #os.makedirs(upload_folder_path, exist_ok=True)

    column_names = list(data.keys())
    print("THESE ARE COLUMN NAMES")
    print(column_names)
    print(type(data))

    # Open the CSV file for writing
    with open(upload_folder_path, 'w', newline='') as file:
        writer = csv.writer(file)

        # Write the header row (keys)
        writer.writerow(data.keys())

        # Write the data row (values)
        writer.writerow(data.values())

def read_user_data(username, upload_folder_path):
    # Create asset and credit file paths
    csv_asset_upload_folder_path = os.path.join(upload_folder_path, f'{username}_assetValue.csv')
    csv_credit_upload_folder_path = os.path.join(upload_folder_path, f'{username}_credits.csv')

    # Read values from asset and credit CSV files
    asset_data = read_csv_file(csv_asset_upload_folder_path)
    credit_data = read_csv_file(csv_credit_upload_folder_path)

    # Create dictionaries for assets and credits
    #assets_dict = {'assets': asset_data}
    #credits_dict = {'credits': credit_data}

    return asset_data, credit_data

def fetch_stock_data(ticker, start_date, end_date):
    yfin.pdr_override()
    #data = pdr.DataReader(ticker, 'yahoo', start_date, end_date).loc[:, 'Adj Close']
    sp = pdr.get_data_yahoo(ticker, start=start_date, end=end_date)
    sp_html = sp.to_html()
    return sp_html, sp

def check_credits():
    try:
        credits = app.config['CREDITS']
    except KeyError:
        return False
    return True

def check_assets():
    try:
        credits = app.config['ASSETS']
    except KeyError:
        return False
    return True

@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('auth/index.html')

@app.route('/dashboard')
@login_required
def dashboard():
    totalCredits = 0
    totalAssets = 0
    if not check_credits() or not check_assets():
        return redirect(url_for('login'))
    else:
        #totalCredits = 0
        for value in app.config['CREDITS'].values():
            totalCredits += int(value)
        #totalAssets = 0
        for value in app.config['ASSETS'].values():
            totalAssets += int(value)

        #labels = [
        #    '2015','2016','2017','2018','2019','2020','2021', '2022','2023'
        #]
        numYears = 10
        labels = []
        labels.append(totalAssets)
        print (numYears*12)
        for i in range(1,((numYears*12)+1)):
            labels.append(i)
        print(len(labels))
        data = []
        assetStart=totalAssets
        #ticker = 'AAPL'  # Example ticker
        ticker = '^SP500TR'
        # Create datetime objects for start and end dates
        #start_date = datetime.datetime(2000, 1, 1)

        #end_date = datetime.datetime(2023, 1, 1)
        # Format datetime objects as strings and assign to original variable names
       # start_date = start_date.strftime("%Y-%m-%d")
        #end_date = end_date.strftime("%Y-%m-%d")
        from dateutil import parser

        # Parse date strings into datetime objects
        start_date = parser.parse("2000-01-01")
        end_date = parser.parse("2023-01-01")

        # Format datetime objects as strings
        start_date = start_date.strftime("%Y-%m-%d")
        end_date = end_date.strftime("%Y-%m-%d")


        stock_data_html, stock_data = fetch_stock_data(ticker, start_date, end_date)
        var1=stock_data.resample('M').last().pct_change().mean().values[0]
        var2=stock_data.resample('M').last().pct_change().std().values[0]
        #print(var1)
        #print(var2)
        np.random.seed(seed=25)
        data.append(assetStart)
        for month in labels:
            market_return = np.random.normal(var1, var2,1)[0]
            assetStart = assetStart * (1+ market_return)
            #print(assetStart)
            data.append(assetStart)
            #print( market_return)
        #print(data)
        
        #data = [0, 10, 15, 8, 22, 18, 25]

        return render_template('pages/dashboard.html', name=current_user.username, total_credits=totalCredits, total_assets=totalAssets, data=data, Labels=labels)

def generate_credits_form(input_list):
    class CreditsForm(DynamicForm):
        pass
    
    for index, field_name in enumerate(input_list):
        initial_value = 0
        field = IntegerField(field_name, validators=[Optional(), NumberRange(min=0)], default=initial_value)
        setattr(CreditsForm, f'field_{index}', field)

    return CreditsForm()

@app.route('/credits', methods=['GET', 'POST'])
@login_required
def credits():
    #input_list = ['value1', 'value2', 'value3']  # Replace with your list
    input_list = app.config['CREDITS'].keys()
    form = generate_credits_form(input_list=input_list)
    #upload_folder_path = os.path.join('Wealth_Managment_Tool/upload_folder', f'{current_user.username}/{current_user.username}_credits.csv')
    #print(upload_folder_path)
    if request.method == 'POST' and form.validate_on_submit():
        old_vals = list(app.config['CREDITS'].values())
        entered_data_list = [value if value != '' else '-1' for value in request.form.values()]
        counter = 0
        for val in entered_data_list:
            if val == '-1' and int(old_vals[counter]) != 0:
                entered_data_list[counter] = old_vals[counter]
            elif val == '-1':
                entered_data_list[counter] = '0'
            counter += 1
        entered_data = {list(input_list)[i]:entered_data_list[i] for i in range(len(input_list))}
        print("Entered data:", entered_data)
        app.config['CREDITS'] = entered_data
        #write_csv_file(upload_folder_path, entered_data)
        write_dict_to_excel(app.config['UPLOAD_FOLDER'], 'credits', entered_data)
        ##print("Entered data:", entered_data)
        flash("Data entered successfully", "success")
        #clearAttribute()

    if request.method == 'POST' and not form.validate_on_submit():
        flash('There was an issue uploading your data, please try again', 'danger')
    
    total_credits = 0
    for key in app.config['CREDITS'].keys():
        total_credits += int(app.config['CREDITS'][key])
    return render_template('pages/credits.html', name=current_user.username, form=form, credits_data = app.config['CREDITS'], total_credits=total_credits)

def generate_assets_form(input_list):
    class AssetsForm(DynamicForm2):
        pass
    
    for index, field_name in enumerate(input_list):
        initial_value = 0
        field = IntegerField(field_name, validators=[Optional(), NumberRange(min=0)], default=initial_value)
        setattr(AssetsForm, f'field_{index}', field)

    return AssetsForm()

@app.route('/assetValue', methods=['GET', 'POST'])
@login_required
def assetValue():
    #input_list = ['value1', 'value2', 'value3']  # Replace with your list
    input_list = app.config['ASSETS'].keys()
    form = generate_assets_form(input_list=input_list)
    #upload_folder_path = os.path.join('Wealth_Managment_Tool/upload_folder', f'{current_user.username}/{current_user.username}_assetValue.csv')
    #print(upload_folder_path)
    if request.method == 'POST' and form.validate_on_submit():
        #print("test 1")
        old_vals = list(app.config['CREDITS'].values())
        entered_data_list = [value if value != '' else '-1' for value in request.form.values()]
        counter = 0
        for val in entered_data_list:
            if val == '-1' and int(old_vals[counter]) != 0:
                entered_data_list[counter] = old_vals[counter]
            elif val == '-1':
                entered_data_list[counter] = '0'
            counter += 1
        entered_data = {list(input_list)[i]:entered_data_list[i] for i in range(len(input_list))}
        #print("test 1")
        app.config['ASSETS'] = entered_data
        #print("test 1")
        #write_csv_file(upload_folder_path, entered_data)
        write_dict_to_excel(app.config['UPLOAD_FOLDER'], 'assets', entered_data)
        #print("test 1")
        #print("Entered data:", entered_data)
        flash("Data entered successfully", "success")

    if request.method == 'POST' and not form.validate_on_submit():
        flash('There was an issue uploading your data, please try again', 'danger')

    total_assets = 0
    for key in input_list:
        total_assets += int(app.config['ASSETS'][key])
    return render_template('pages/currentAssetValue.html', name=current_user.username, form=form, assets_data=app.config['ASSETS'], total_assets=total_assets)

@app.route('/simulatedGrowth')
@login_required
def simulatedGrowth():
    totalAssets = 0
    for value in app.config['ASSETS'].values():
            totalAssets += int(value)

    numYears = 10
    labels = []
    labels.append(totalAssets)
    print (numYears*12)
    for i in range(1,((numYears*12)+1)):
        labels.append(i)
    print(len(labels))
    SPdata = [] #S+P500 data
    Sdata = [] #savings data
    LRdata =[] #Low risk managed
    HRdata =[] #Hish Risk managed

    assetStart=totalAssets
    #ticker = 'AAPL'  # Example ticker
    ticker = '^SP500TR'
    from dateutil import parser

    # Parse date strings into datetime objects
    start_date = parser.parse("2000-01-01")
    end_date = parser.parse("2023-01-01")

    # Format datetime objects as strings
    start_date = start_date.strftime("%Y-%m-%d")
    end_date = end_date.strftime("%Y-%m-%d")


    stock_data_html, stock_data = fetch_stock_data(ticker, start_date, end_date)
    var1=stock_data.resample('M').last().pct_change().mean().values[0]
    var2=stock_data.resample('M').last().pct_change().std().values[0]
    np.random.seed(seed=25)
    SPdata.append(assetStart)
    LRdata.append(assetStart)
    Sdata.append(assetStart)
    HRdata.append(assetStart)
    SP_asset_val = assetStart
    LR_asset_val = assetStart
    S_asset_val = assetStart
    HR_asset_val = assetStart
    for i in  range(len(labels)):
        market_return = np.random.normal(var1, var2,1)[0]
        #SPreturn = assetStart * (1+ market_return)
        SP_asset_val *= (1+ market_return)
        SPdata.append(SP_asset_val)
        LR_asset_val *= (1 + ((np.random.normal(0.65, 0.35,1)[0]/12)/100))
        LRdata.append(LR_asset_val)
        if i < 12:
            S_asset_val *= (1 + ((np.random.normal(7, 4,1)[0]/12)/100))
        else:
            S_asset_val *= (1 + ((np.random.normal(2, 2,1)[0]/12)/100))
        Sdata.append(S_asset_val)
        HR_asset_val *= (1+ ((np.random.normal(7, 20,1)[0]/12)/100))
        HRdata.append(HR_asset_val)
    return render_template('pages/simulatedGrowth.html', name=current_user.username, stock_table = stock_data_html, labels=labels, SPdata=SPdata, Sdata=Sdata, LRdata=LRdata, HRdata=HRdata)

@app.route('/tools', methods=['GET', 'POST'])
@login_required
def tools():
    return render_template('pages/tools.html')


@app.route('/updateFinances')
@login_required
def updateFinances():
    return render_template('pages/updateFinancialDetails.html', name=current_user.username)

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()

    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user:
            #intial issues with the password hash function when deployed
            #flash("User found Password"+user.password, "danger")
            #if check_password_hash(user.password, form.password.data):
            if user.password == form.password.data:
                login_user(user, remember=form.remember.data)
                username = current_user.username
                upload_folder_path = os.path.join('Wealth_Managment_Tool/upload_folder', username, f'{username}_values.xlsx')

                # Set the UPLOAD_FOLDER configuration
                app.config['UPLOAD_FOLDER'] = upload_folder_path
                #assets, credits = read_user_data(username, upload_folder_path)
                listOfDictionaries = excel_to_dict(app.config['UPLOAD_FOLDER'])
                app.config['ASSETS'] = listOfDictionaries['assets'][0]
                app.config['CREDITS'] = listOfDictionaries['credits'][0]
                app.config['RETIREMENT'] = listOfDictionaries['retirement'][0]
                app.config['SAVINGS'] = listOfDictionaries['savings'][0]
                app.config['FINANCE_PATH'] = os.path.join('Wealth_Managment_Tool/SimulatedFinanceData/', username)
                #print(assets)
                #print(credits)
                flash("Welcome "+ str(current_user.username)+", you have been logged in.")
                return redirect(url_for('dashboard'))

        flash("Incorrect Username or Password", "danger")
        #return '<h1>Invalid Username or Password</h1>'
        #return '<h1>' + form.username.data + ' ' + form.password.data + '</h1>'

    return  render_template('login/login.html',form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegistrationForm()
    
    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data)
        #new_user = User(username=form.username.data, email=form.email.data, password=hashed_password)
        new_user = User(username=form.username.data, email=form.email.data, password=form.password.data)
        db.session.add(new_user)
        db.session.commit()
        username = form.username.data
        upload_folder_path = os.path.join('Wealth_Managment_Tool/upload_folder', username)
        #csv_asset_upload_folder_path = os.path.join(upload_folder_path, f'{username}_assetValue.csv')
        #csv_credit_upload_folder_path = os.path.join(upload_folder_path, f'{username}_credits.csv')

        # Ensure the folder exists (create it if it doesn't)
        os.makedirs(upload_folder_path, exist_ok=True)
        upload_folder_path = os.path.join(upload_folder_path, f'{username}_values.xlsx')
        # Set the UPLOAD_FOLDER configuration
        app.config['UPLOAD_FOLDER'] = upload_folder_path
        #Create a user csv which will store their asset values
        #with open(csv_asset_upload_folder_path, 'w', newline='') as file:
        #    writer = csv.writer(file)
        #    writer.writerow(['House', 'Car', 'Investments', 'Checking Account', 'Stocks', 'Savings', 'Retirement Accounts'])
        #    writer.writerow([])

        #with open(csv_credit_upload_folder_path, 'w', newline='') as file:
        #    writer = csv.writer(file)
        #    writer.writerow(['Rent', 'Mortgage', 'Utilities', 'Food and Groceries','Car Payments', 'Student loan Payments', 'Pension','Streaming Subscriptions', 'Music Subscriptions', 'Misc Subscriptions', 'Health Insurance', 'House Insurance', 'Other Insurance'])
        #    writer.writerow([20, 400])

        # Define column names for asset and credit files
        asset_column_names = ['House', 'Car', 'Investments', 'Checking Account', 'Stocks', 'Savings', 'Retirement Accounts']
        credit_column_names = ['Rent', 'Mortgage', 'Utilities', 'Food and Groceries', 'Car Payments', 'Student loan Payments', 'Pension', 'Subscriptions', 'Health Insurance', 'House Insurance', 'Other Insurance']

        sheet_data = {
            'credits': credit_column_names,  # List for 'credits' sheet
            'assets': asset_column_names,      # List for 'assets' sheet
            'retirement': ['Retirement Age'],  # List for 'retirement' sheet
            'savings': ['Saving Goal'] # List for 'savings' sheet
        }
        # Create dictionaries with column names as keys and 0 as values
        #asset_data_dict = dict.fromkeys(asset_column_names, 0)
        #credit_data_dict = dict.fromkeys(credit_column_names, 0)

        # Create asset and credit files with dictionaries
        #create_csv_file(csv_asset_upload_folder_path, asset_column_names, asset_data_dict)
        #create_csv_file(csv_credit_upload_folder_path, credit_column_names, credit_data_dict)
        create_excel_file(app.config['UPLOAD_FOLDER'], sheet_data=sheet_data)
        finance_path = 'Wealth_Managment_Tool/SimulatedFinanceData/' + username
        os.makedirs(finance_path, exist_ok=True)

        generator = FinanceDataGenerator()
        generator.generate_and_save_json(finance_path + '/current_account_transactions.json')

        generator = CreditCardDataGenerator()
        generator.generate_and_save_json(finance_path + '/credit_card_transactions.json')
        app.config['FINANCE_PATH'] = finance_path

        flash('New User Created', 'success')
        return redirect(url_for('login'))
        #return '<h1>' + form.email.data + ' ' + form.username.data + ' ' + form.password.data + '</h1>'

    return render_template('login/register.html',form=form)

@app.route('/upload',  methods=['GET', 'POST'])
@login_required
def upload():
    form = UploadForm()
    if form.validate_on_submit():
        file = form.file.data
        filename = file.filename
        file.save('Wealth_Managment_Tool/upload_folder', current_user.username, 'banking_data' + filename)
        flash('File uploaded successfully', 'success')
        return redirect('/upload')  # Redirect to the same page after successful upload
    return render_template('pages/uploadForm.html', form=form)

@app.route('/transactions')
@login_required
def transactions():
    # Read data from JSON files for both account and credit card

    accounts_path = os.path.join(app.config['FINANCE_PATH'],'current_account_transactions.json')
    print (accounts_path)
    credits_path = os.path.join(app.config['FINANCE_PATH'], 'credit_card_transactions.json')
    print (credits_path)

    with open(accounts_path, 'r') as f:
        account_data = json.load(f)

    with open(credits_path, 'r') as f:
        credit_card_data = json.load(f)

    # Extract account information
    account_info = account_data['account']
    account_transactions_data = account_data['transactions']

    # Extract credit card information
    credit_card_info = credit_card_data['account']
    credit_card_transactions_data = credit_card_data['transactions']

    # Instantiate Account object
    account = Account(
        account_number=account_info['account_number'],
        sort_code=account_info['sort_code'],
        balance=account_info['balance'],
        currency=account_info['currency'],
        account_holder=account_info['account_holder'],
        transactions=[Transaction(**txn_data) for txn_data in account_transactions_data],
    )

    # Instantiate Credit Card object
    credit_card = CreditCard(
        card_number=credit_card_info['card_number'],
        expiry_date=credit_card_info['expiry_date'],
        card_holder=credit_card_info['card_holder'],
        credit_limit=credit_card_info['credit_limit'],
        currency=credit_card_info['currency'],
        transactions=[Transaction(**txn_data) for txn_data in credit_card_transactions_data],
    )

    # Convert account and credit card objects to dictionaries
    account_dict = account.to_dict()
    credit_card_dict = credit_card.to_dict()
    account_balances = [transaction.balance for transaction in account.transactions] 
    print(account_balances)
    balances_labels = []
    counter = 1
    for transaction in account_balances:
        balances_labels.append(counter)
        counter += 1
    credit_balances = [transaction.balance for transaction in credit_card.transactions]
    print(credit_balances)
    

    # Render HTML template with both account and credit card information
    return render_template('pages/transactions.html', current_account=account_dict, credit_card=credit_card_dict, name=current_user.username, account_balances=account_balances, credit_balances=credit_balances, balances_labels=balances_labels)

@app.route('/delete_files', methods=['GET', 'POST'])
@login_required
def delete_files():
    form = DeleteFileForm()

    # Specify the directory path you want to iterate through
    directory_path = app.config['UPLOAD_FOLDER']

    file_names = list_files(directory_path)

    # Populate choices for the SelectMultipleField
    form.files_to_delete.choices = [(file, file) for file in file_names]

    if form.validate_on_submit():
        # Get the selected file names and perform deletion logic
        selected_files = form.files_to_delete.data
        for file in selected_files:
            file_path = os.path.join(directory_path, file)
            # Add logic to delete the file or perform any other actions
            try:
                os.remove(file_path)
            except OSError as e:
                file_path = ''

        return redirect(url_for('dashboard'))

    return render_template('pages/deleteFiles.html', form=form)

@app.route('/double_form', methods=['GET', 'POST'])
@login_required
def double_form():
    form1 = Form1(request.form)
    form2 = Form2(request.form)
    card_content1 = None
    card_content2 = None

    if request.method == 'POST':
        if form1.validate():
            field1_value = form1.field1.data
            card_content1 = f'New savings goal: {field1_value}'
            app.config['RETIREMENT']['Retirement Age'] = field1_value
            write_dict_to_excel(app.config['UPLOAD_FOLDER'], 'retirement', app.config['RETIREMENT'])
            old_savings = app.config['SAVINGS']['Saving Goal']
            card_content2 = f'Previous savings goal: {old_savings}'
        elif form2.validate():
            field2_value = form2.field2.data
            card_content2 = f'New desired retirement age: {field2_value}'
            app.config['SAVINGS']['Saving Goal'] = field2_value
            write_dict_to_excel(app.config['UPLOAD_FOLDER'], 'savings', app.config['SAVINGS'])
            old_retirement = app.config['RETIREMENT']['Retirement Age']
            card_content1 = f'Previous retirement goal: {old_retirement}'

    ### INSERT FUNCTION TO SAVE THE FORM INPUTS, AND SET THE OPPOSITE CARD TO THE SUBMISSION OF PREVIOUSLY SAVED IF APPLICABLE
    
    return render_template('pages/doubleForm.html', form1=form1, form2=form2, card_content1=card_content1, card_content2=card_content2)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))









@app.route('/graph')
def graph():
    df = pd.read_excel('example.xlsx')
    cols = df[['rank', 'total_assets_us_b']]

    out = cols.to_numpy().tolist()
    out = [tuple(elt) for elt in out]
    out = [("sdiufnd", 23213),
    ("wksdfsdf", 123123),
    ("iwe", 2312)]

    labels = [row[0] for row in out]
    values = [row[1] for row in out]
    # Render the HTML template with the base64-encoded image
    #table_html = df.to_html(classes='table table-striped', index=False)
    return render_template('viewData/graph.html', labels=labels, values=values)

#@app.route('/monzo_test')
#def monzo():
#    return render_template('pages/monzo_test.html')

@app.route('/capitalOne', methods=['GET', 'POST'])
def capitalOne():
    # Define the endpoint and request body
    api_url = 'https://api-sandbox.capitalone.com/oauth2/token'
    request_body = {
        'client_id': '9c08b0d10faec81a311401d34e99ccf7',
        'client_secret': '48af6fca7486763d5982ee1e5ff7bb99',
        'grant_type': 'client_credentials'
    }

    # Make the POST request
    response = requests.post(api_url, data=request_body, headers={'Content-Type': 'application/x-www-form-urlencoded'})

    # Check if the request was successful
    if response.status_code == 200:
        token_data = response.json()
        access_token = token_data['access_token']
        print('Access Token:', access_token)
    else:
        print('Error:', response.text)
    

    api_url = 'https://api-sandbox.capitalone.com/deposits/products/~/search'
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json;v=5',
        'Accept': 'application/json;v=5'
    }
    body = {
        'isCollapseRate': True
    }
    response = requests.post(api_url, headers=headers, json=body)
    if response.status_code == 200:
        print('Request successful')
        print(response.json())
    else:
        print('Error:', response.text)


    return(redirect(url_for('dashboard')))


@app.route('/monzo', methods=['GET', 'POST'])
def monzo():
    client_id = 'oauth2client_0000AdfM5AeMUPFJkXTHnt'
    redirect_uri = 'http://localhost'
    state_token = 'your_state_token'

    url = f"https://auth.monzo.com/?client_id={client_id}&redirect_uri={redirect_uri}&response_type=code&state={state_token}"
    
    return(redirect(url))

@app.route('/oauth/callback')
def oauth_callback():
    # Get the authorization code and state token from the URL parameters
    #authorization_code = request.args.get('code')
    #state_token = request.args.get('state')

    # Check if state_token matches the one you provided earlier
    #expected_state_token = 'your_state_token'

    #if state_token != expected_state_token:
    #    return 'Invalid state token', 400

    # Process the authorization code, exchange it for an access token, etc.
    # You will need to implement this part according to Monzo's documentation

    # For now, let's just print out the received authorization code
    #print(f'Received authorization code: {authorization_code}')

    # You can also redirect the user to another page if needed
    # return redirect('/success')

    # Return a success message
    #return 'Authorization successful'
    authorization_code = 'eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCJ9.eyJlYiI6InZmWEl1VkZxSzkrSktIR3Y0WXlrIiwianRpIjoiYXV0aHpjb2RlXzAwMDBBZXdraFlDSjhUZGdEdmdKWE4iLCJ0eXAiOiJhemMiLCJ2IjoiNiJ9.5rWu7Vr4aZ8qaGvsVAHeNRV5PvsKndDueIGb6kaaf4yi0BNzcCYTI27ntpbHW3c9R1XnKqQz3XEyxxi4gJhciw'

    access_token(authorization_code)
    return(redirect(url_for('login')))

def access_token(authorization_code):
    params = {
        'grant_type': 'authorization_code',
        'client_id': 'oauth2client_0000AdfM5AeMUPFJkXTHnt',
        'client_secret': 'mnzconf.libMV5WRpToEiz+Q64aUwZCWMLUcr110p4OSgvEY5TwSxT+SoX0xcsM7fC0fciBUrBYZIkCFLj4akwpF2qpkwQ==',
        'redirect_uri': 'http://localhost',
        'code': authorization_code
    }
    try:
        response = requests.post("https://api.monzo.com/oauth2/token", data=params)
        response.raise_for_status()  # Raise an exception for non-200 status codes
        access_token = response.json().get('access_token')
        if access_token:
            response = requests.get("https://api.monzo.com/balance?account_id=user_00009hPDVBoQ6BJJXislbV", headers={"Authorization": f"Bearer {access_token}"})
            response.raise_for_status()  # Raise an exception for non-200 status codes
            return jsonify(response.json())
        else:
            return jsonify({"error": "Access token not found"}), 500
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 500
    

#with app.app_context():
    #db.drop_all()
    #db.create_all()
    #sample_user = User(username="test", email="test@test.com", password=generate_password_hash("password"))
    #db.session.add(sample_user)
    #db.session.commit()
    
    #users = User.query.all()
    #print("Print Statement")
    #print(users)
    #print("End print statement")

if __name__ == '__main__':
    app.run(debug=True)